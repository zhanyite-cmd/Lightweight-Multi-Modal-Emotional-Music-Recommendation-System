import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union, Callable
import os
import io
import base64
import json
import textwrap

import cv2
import numpy as np
from PIL import Image, ImageTk
import shutil
try:
    import winsound  # Windows alarm
except Exception:  # pragma: no cover
    winsound = None

try:
    from ultralytics import YOLO
except Exception as exc:  # pragma: no cover
    YOLO = None
    _import_error = exc
else:
    _import_error = None

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import tkinter.font as tkfont
import tkinter.font as tkfont


def _iou(box_a: List[int], box_b: List[int]) -> float:
    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b
    inter_x1 = max(ax1, bx1)
    inter_y1 = max(ay1, by1)
    inter_x2 = min(ax2, bx2)
    inter_y2 = min(ay2, by2)
    inter_w = max(0, inter_x2 - inter_x1)
    inter_h = max(0, inter_y2 - inter_y1)
    inter = inter_w * inter_h
    area_a = max(0, ax2 - ax1) * max(0, ay2 - ay1)
    area_b = max(0, bx2 - bx1) * max(0, by2 - by1)
    union = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


class SimpleTracker:
    def __init__(self, max_age: int = 30, iou_threshold: float = 0.3) -> None:
        self.max_age = max_age
        self.iou_threshold = iou_threshold
        self.next_id = 1
        self.tracks: Dict[int, Dict] = {}

    def update(self, detections: List[Dict]) -> List[Dict]:
        # Increment ages
        for tr in self.tracks.values():
            tr["age"] += 1

        assigned_dets = set()
        # Try to match existing tracks
        for tid, tr in list(self.tracks.items()):
            best_iou = 0.0
            best_idx = -1
            for idx, det in enumerate(detections):
                if idx in assigned_dets:
                    continue
                i = _iou(tr["box"], det["box"])
                if i > best_iou:
                    best_iou = i
                    best_idx = idx
            if best_iou >= self.iou_threshold and best_idx >= 0:
                det = detections[best_idx]
                tr["box"] = det["box"]
                tr["conf"] = det.get("conf", 0.0)
                tr["age"] = 0
                assigned_dets.add(best_idx)

        # Create tracks for unmatched detections
        for idx, det in enumerate(detections):
            if idx in assigned_dets:
                continue
            tid = self.next_id
            self.next_id += 1
            self.tracks[tid] = {"id": tid, "box": det["box"], "conf": det.get("conf", 0.0), "age": 0}

        # Remove stale tracks
        for tid in list(self.tracks.keys()):
            if self.tracks[tid]["age"] > self.max_age:
                del self.tracks[tid]

        return list(self.tracks.values())


class FaceDetectionApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("A face detection and expression analysis system based on deep learning")
        self.root.geometry("1200x720")

        # Menu bar
        self._init_menu_bar()

        # Model & inference state
        self.model = None
        self.cls_model = None
        self.weights_path = None  # type: Optional[Path]
        self.cls_weights_path = None  # type: Optional[Path]
        self.conf_threshold = tk.DoubleVar(value=0.25)
        self.camera_index = tk.IntVar(value=0)
        self.is_running: bool = False
        self.capture = None  # type: Optional[cv2.VideoCapture]
        self.worker_thread = None  # type: Optional[threading.Thread]
        self.frame_lock = threading.Lock()
        self.last_frame_bgr = None  # type: Optional[np.ndarray]
        self.stop_event = threading.Event()

        # Metrics
        self.current_fps: tk.StringVar = tk.StringVar(value="FPS: -")
        self.status_text: tk.StringVar = tk.StringVar(value="Model not loaded.")
        self.time_info_var = tk.StringVar(value="Time used: - ms")
        self.thresh_info_var = tk.StringVar(value="Threshold: 0.25")
        self.target_info_var = tk.StringVar(value="Target number: 0")
        self.thresh_value_var = tk.StringVar(value="0.25")

        # Output directory
        self.output_dir = Path("runs/face_gui")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # History & export
        self.history = []  # type: List[Dict]
        self.history_counter = 1
        # Counter for detection table rows (for video cumulative display)
        self.detection_counter = 1

        # Alarm
        self.alarm_enabled: tk.BooleanVar = tk.BooleanVar(value=False)
        self.alarm_min_faces: tk.IntVar = tk.IntVar(value=1)
        self._last_alarm_ts: float = 0.0

        # Optional: record live detections to history
        self.record_history_enabled: tk.BooleanVar = tk.BooleanVar(value=False)
        self._last_hist_ts: float = 0.0

        # Tracker disabled (no ID feature)
        self.tracker = None

        # Save-state for on-demand saving
        self.last_annotated_image = None  # type: Optional[np.ndarray]
        self.last_batch_dir = None  # type: Optional[Path]
        self.last_video_path = None  # type: Optional[str]

        self._build_ui()

    def _build_ui(self) -> None:
        # Header title at top center (larger, bold)
        try:
            _big_font = tkfont.nametofont("TkDefaultFont").copy()
            try:
                _base_size = int(_big_font.cget("size"))
            except Exception:
                _base_size = 10
            _big_font.configure(size=_base_size + 6, weight="bold")
            header = tk.Label(self.root, text="A face detection and expression analysis system based on deep learning", font=_big_font, fg="white", bg="#1976D2"  )
        except Exception:
            header = tk.Label(self.root, text="A face detection and expression analysis system based on deep learning", bg="#1976D2", fg="white")
        header.pack(side=tk.TOP, fill=tk.X, padx=8, pady=(10, 4))

        # Top controls frame
        controls = ttk.Frame(self.root)
        controls.pack(side=tk.TOP, fill=tk.X, padx=8, pady=6)

        # Bold style for important section headers
        try:
            style = ttk.Style()
            base_font = tkfont.nametofont("TkDefaultFont").copy()
            try:
                base_size = int(base_font.cget("size"))
            except Exception:
                base_size = 10
            base_font.configure(weight="bold", size=base_size + 2)
            style.configure("Bold.TLabelframe.Label", font=base_font)
        except Exception:
            pass

        # Group weight button and its label vertically so the label sits under the button
        weights_col = ttk.Frame(controls)
        weights_col.pack(side=tk.LEFT, padx=(0, 8))
        style.configure("Green.TButton", background="#4CAF50", foreground="black")
        style.configure("Blue.TButton", background="#2196F3", foreground="black")
        ttk.Button(weights_col, text="Select facial weight", style="Green.TButton",command=self.on_select_weights).pack(side=tk.TOP, anchor=tk.W)
        self.weights_label_var = tk.StringVar(value="Face weight: Not selected")
        self.weights_label = ttk.Label(weights_col, textvariable=self.weights_label_var, wraplength=260, anchor=tk.W, justify=tk.LEFT)
        self.weights_label.pack(side=tk.TOP, fill=tk.X)

        # Classification weights selector
        ttk.Button(weights_col, text="Select expression weights",style="Blue.TButton", command=self.on_select_cls_weights).pack(side=tk.TOP, anchor=tk.W, pady=(6, 0))
        self.cls_weights_label_var = tk.StringVar(value="Expression weight: Not selected")
        self.cls_weights_label = ttk.Label(weights_col, textvariable=self.cls_weights_label_var, wraplength=260, anchor=tk.W, justify=tk.LEFT)
        self.cls_weights_label.pack(side=tk.TOP, fill=tk.X)

        ttk.Label(controls, text="threshold value").pack(side=tk.LEFT, padx=(16, 4))
        self.threshold_scale = ttk.Scale(
            controls, from_=0.05, to=0.95, orient=tk.HORIZONTAL,
            variable=self.conf_threshold, command=lambda _=None: self._on_threshold_change()
        )
        self.threshold_scale.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        ttk.Label(controls, textvariable=self.thresh_value_var, width=5).pack(side=tk.LEFT, padx=(0, 8))

        ttk.Label(controls, textvariable=self.current_fps).pack(side=tk.LEFT, padx=(8, 8))

        # Camera controls grouped together with label
        cam_controls = ttk.LabelFrame(controls, text="real-time detection", style="Bold.TLabelframe")
        cam_controls.pack(side=tk.LEFT, padx=(8, 8), pady=(0, 0))

        ttk.Label(cam_controls, text="camera ID").pack(side=tk.LEFT, padx=(0, 4))
        self.cam_entry = ttk.Entry(cam_controls, textvariable=self.camera_index, width=4)
        self.cam_entry.pack(side=tk.LEFT)

        style.configure("Green.TButton",  background="#4CAF50", foreground="black")   # 亮绿 -
        style.configure("Blue.TButton",   background="#2196F3", foreground="black")   # 天蓝 -
        style.configure("Red.TButton",    background="#F44336", foreground="black")   # 鲜红 -
        style.configure("Orange.TButton", background="#FF9800", foreground="black")   # 橙色 -
        style.configure("Purple.TButton", background="#9C27B0", foreground="black")   # 紫色 -
        style.configure("Teal.TButton",   background="#009688", foreground="black")   # 蓝绿色
        style.configure("Gray.TButton",   background="#9E9E9E", foreground="black")   # 中灰 -
        style.configure("Yellow.TButton", background="#FFEB3B", foreground="black")   # 明黄 -
        style.configure("DarkBlue.TButton", background="#3F51B5", foreground="black") # 深蓝 -
        style.configure("Pink.TButton",   background="#E91E63", foreground="black")   # 粉红 -
        self.start_btn = ttk.Button(cam_controls, text="Start the camera", style="Green.TButton", command=self.on_start_camera)
        self.start_btn.pack(side=tk.LEFT, padx=(8, 4))

        self.stop_btn = ttk.Button(cam_controls, text="stop", style="Red.TButton", command=self.on_stop_camera, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=(4, 4))

        self.save_btn = ttk.Button(cam_controls, text="Save the current frame", style="Orange.TButton", command=self.on_save_frame)
        self.save_btn.pack(side=tk.LEFT, padx=(8, 0))

        # File-based detection controls grouped together with label
        file_controls = ttk.LabelFrame(controls, text="Image/Video/Batch Inspection", style="Bold.TLabelframe")
        file_controls.pack(side=tk.LEFT, padx=(8, 8))
        # Left group: action buttons arranged 2x2 (two per row)
        fc_left = ttk.Frame(file_controls)
        fc_left.pack(side=tk.LEFT)
        row1 = ttk.Frame(fc_left)
        row1.pack(side=tk.TOP, fill=tk.X)
        row2 = ttk.Frame(fc_left)
        row2.pack(side=tk.TOP, fill=tk.X)
        self.btn_open_image = ttk.Button(row1, text="打开图片检测", style="Blue.TButton", command=self.on_detect_image)
        self.btn_open_image.pack(side=tk.LEFT, padx=(4, 4), pady=(2, 2))
        self.btn_photo_capture = ttk.Button(row1, text="拍照检测", style="Purple.TButton", command=self.on_capture_photo_detect)
        self.btn_photo_capture.pack(side=tk.LEFT, padx=(4, 4), pady=(2, 2))
        self.btn_select_video = ttk.Button(row2, text="选择视频检测", style="Teal.TButton", command=self.on_select_video_detect)
        self.btn_select_video.pack(side=tk.LEFT, padx=(4, 4), pady=(2, 2))
        self.btn_batch = ttk.Button(row2, text="批量图片检测", style="Gray.TButton", command=self.on_batch_images_detect)
        self.btn_batch.pack(side=tk.LEFT, padx=(4, 8), pady=(2, 2))
        # Right-side vertical action column: 保存检测结果 (top) and 导出结果 (bottom)
        right_actions = ttk.Frame(controls)
        right_actions.pack(side=tk.RIGHT, padx=(4, 8))
        self.save_result_btn = ttk.Button(right_actions, text="保存检测结果", style="Yellow.TButton", command=self.on_save_detection_result, state=tk.DISABLED)
        self.save_result_btn.pack(side=tk.TOP, fill=tk.X, pady=(0, 4))

        # Keep for enabling/disabling during video processing
        self._file_buttons = [self.btn_open_image, self.btn_photo_capture, self.btn_select_video, self.btn_batch, self.save_result_btn]

        # Alarm controls moved into camera group
        ttk.Checkbutton(cam_controls, text="报警", variable=self.alarm_enabled).pack(side=tk.LEFT, padx=(8, 4))
        ttk.Label(cam_controls, text="最小人数").pack(side=tk.LEFT, padx=(4, 4))
        self.alarm_spin = ttk.Spinbox(cam_controls, from_=1, to=99, textvariable=self.alarm_min_faces, width=3)
        self.alarm_spin.pack(side=tk.LEFT, padx=(0, 8))

        ttk.Checkbutton(cam_controls, text="记录到历史", variable=self.record_history_enabled).pack(side=tk.LEFT, padx=(4, 8))

        ttk.Button(right_actions, text="导出历史记录", style="DarkBlue.TButton", command=self.on_export_results).pack(side=tk.TOP, fill=tk.X, pady=(0, 4))
        ttk.Button(right_actions, text="导出检测信息", style="Pink.TButton", command=self.on_export_detection_info).pack(side=tk.TOP, fill=tk.X)

        # Main body: canvas (left) + info panel (right)
        body = ttk.Frame(self.root)
        body.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Canvas for video/image display - fixed size to prevent resizing with video dimensions
        # Use a Frame wrapper to set fixed pixel dimensions
        canvas_frame = tk.Frame(body, width=900, height=440, bg="#111111")
        canvas_frame.pack(side=tk.LEFT, padx=(0, 4))  # Reduced padding to minimize gap
        canvas_frame.pack_propagate(False)  # Prevent frame from resizing based on content
        self.canvas = tk.Label(canvas_frame, bg="#111111")
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # Info panel on the right - fixed width to prevent being squeezed
        # Width set to 530 to accommodate all table columns (7 columns: idx 60 + emotion 70 + emotion_conf 90 + x1/y1/x2/y2 60*4 + scrollbar + padding)
        info_frame = ttk.LabelFrame(body, text="Detection information", style="Bold.TLabelframe", width=530)
        info_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(4, 0))  # Reduced padding to minimize gap
        info_frame.pack_propagate(False)
        ttk.Label(info_frame, textvariable=self.current_fps).pack(anchor=tk.W, padx=8, pady=(6, 2))
        ttk.Label(info_frame, textvariable=self.time_info_var).pack(anchor=tk.W, padx=8, pady=2)
        ttk.Label(info_frame, textvariable=self.target_info_var).pack(anchor=tk.W, padx=8, pady=2)
        ttk.Label(info_frame, textvariable=self.thresh_info_var).pack(anchor=tk.W, padx=8, pady=(2, 8))

        cols = ("idx", "emotion", "emotion_conf", "x1", "y1", "x2", "y2")
        self.dets_tv = ttk.Treeview(info_frame, columns=cols, show="headings", height=14)
        headers = [("idx", "oder", 60), ("emotion", "Emotion", 70), ("emotion_conf", "Expression confidence level", 90), ("x1", "x1", 60), ("y1", "y1", 60), ("x2", "x2", 60), ("y2", "y2", 60)]
        for key, txt, w in headers:
            self.dets_tv.heading(key, text=txt)
            self.dets_tv.column(key, width=w, anchor=tk.W)
        self.dets_tv.pack(side=tk.LEFT, fill=tk.Y, padx=(8, 0), pady=(0, 8))
        sb_tv = ttk.Scrollbar(info_frame, orient=tk.VERTICAL, command=self.dets_tv.yview)
        self.dets_tv.configure(yscrollcommand=sb_tv.set)
        sb_tv.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 8))

        # History panel - fixed height to prevent being squeezed
        # Increased height to 240 for better visibility
        history_frame = ttk.LabelFrame(self.root, text="Historical record", style="Bold.TLabelframe", height=240)
        history_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=8, pady=(0, 8))
        history_frame.pack_propagate(False)
        columns = ("idx", "path", "emotion", "emotion_conf", "bbox", "time")
        self.history_tv = ttk.Treeview(history_frame, columns=columns, show="headings", height=6)
        headers = (
            ("idx", "oder", 60),
            ("path", "file path", 420),
            ("emotion", "emotion", 120),
            ("emotion_conf", "Expression confidence level", 80),
            ("bbox", "target location", 260),
            ("time", "time", 160),
        )
        for col, txt, w in headers:
            self.history_tv.heading(col, text=txt)
            self.history_tv.column(col, width=w, anchor=tk.W)
        self.history_tv.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sb = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=self.history_tv.yview)
        self.history_tv.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        # Status bar
        status_bar = ttk.Frame(self.root)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Label(status_bar, textvariable=self.status_text).pack(side=tk.LEFT, padx=8, pady=4)

        # Keyboard shortcuts
        self.root.bind("<Escape>", lambda _evt: self.on_stop_camera())

    def _init_menu_bar(self) -> None:
        try:
            menubar = tk.Menu(self.root)
            func_menu = tk.Menu(menubar, tearoff=0)
            # func_menu.add_command(label="人脸检测", command=lambda: self._switch_to_face_detection())
            func_menu.add_command(label="表情分析", command=lambda: self._open_emotion_analysis())
            menubar.add_cascade(label="功能", menu=func_menu)
            # Help menu
            help_menu = tk.Menu(menubar, tearoff=0)
            help_menu.add_command(label="关于", command=lambda: messagebox.showinfo("about", "Face Detection and Expression Analysis System"))
            help_menu.add_separator()
            help_menu.add_command(label="退出", command=self.root.destroy)
            menubar.add_cascade(label="帮助", menu=help_menu)
            self.root.config(menu=menubar)
            self.menubar = menubar
        except Exception:
            pass

    def _switch_to_face_detection(self) -> None:
        try:
            # If emotion window exists, close it and show main
            if getattr(self, "_emotion_win", None) is not None:
                try:
                    self._emotion_win.destroy()
                except Exception:
                    pass
                self._emotion_win = None
            self.root.deiconify()
            self.root.lift()
        except Exception:
            pass

    def _open_emotion_analysis(self) -> None:
        # Hide main window and open emotion analysis window
        try:
            self.root.withdraw()
        except Exception:
            pass
        try:
            self._emotion_win = EmotionAnalysisWindow(parent_app=self)
        except Exception:
            # If creation fails, restore main window
            try:
                self.root.deiconify()
            except Exception:
                pass
            raise

    # ===================== Event Handlers =====================
    def _on_threshold_change(self) -> None:
        # Live threshold is taken each inference
        try:
            val = float(self.conf_threshold.get())
            self.thresh_info_var.set(f"阈值: {val:.2f}")
            self.thresh_value_var.set(f"{val:.2f}")
        except Exception:
            self.thresh_info_var.set("阈值: -")
            self.thresh_value_var.set("")

    def on_select_weights(self) -> None:
        path = filedialog.askopenfilename(
            title="Select the YOLO weight file",
            filetypes=[("PyTorch Weights", "*.pt"), ("All Files", "*.*")],
            initialdir=str(Path.cwd())
        )
        if not path:
            return
        self.weights_path = Path(path)
        self._update_weights_display()
        # Model label stays only on top bar; info panel doesn't show model now
        self._load_model_async()

    def on_select_cls_weights(self) -> None:
        path = filedialog.askopenfilename(
            title="Select the YOLO classification weight file",
            filetypes=[("PyTorch Weights", "*.pt"), ("All Files", "*.*")],
            initialdir=str(Path.cwd())
        )
        if not path:
            return
        self.cls_weights_path = Path(path)
        self._update_cls_weights_display()
        self._load_cls_model_async()

    def on_start_camera(self) -> None:
        if self.model is None:
            messagebox.showwarning("提示", "Please first select and load the detection model weights!")
            return
        if self.is_running:
            return
        index = int(self.camera_index.get())
        self.capture = cv2.VideoCapture(index, cv2.CAP_DSHOW)
        if not self.capture.isOpened():
            messagebox.showerror("错误", f"The camera cannot be opened. {index}")
            return
        # Reset detection counter and clear table for new camera session (camera: cumulative mode)
        self.detection_counter = 1
        self.dets_tv.delete(*self.dets_tv.get_children())
        self.is_running = True
        self.stop_event.clear()
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.status_text.set("Real-time detection in progress... Press Esc to stop.")
        self.worker_thread = threading.Thread(target=self._camera_loop, daemon=True)
        self.worker_thread.start()

    def on_stop_camera(self) -> None:
        if not self.is_running:
            return
        self.is_running = False
        self.stop_event.set()
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status_text.set("stop")
        if self.capture is not None:
            try:
                self.capture.release()
            except Exception:
                pass
            self.capture = None

    def on_save_frame(self) -> None:
        with self.frame_lock:
            frame = None if self.last_frame_bgr is None else self.last_frame_bgr.copy()
        if frame is None:
            messagebox.showinfo("提示", "无可保存帧。请先开始检测。")
            return
        # Ask user where to save the snapshot
        default_name = f"snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        path = filedialog.asksaveasfilename(
            title="保存当前帧为…",
            defaultextension=".jpg",
            initialfile=default_name,
            filetypes=[("JPEG", "*.jpg"), ("PNG", "*.png"), ("All Files", "*.*")]
        )
        if not path:
            return
        try:
            cv2.imwrite(str(path), frame)
            self.status_text.set(f"已保存: {path}")
        except Exception as exc:
            messagebox.showerror("错误", f"保存失败: {exc}")
            return

    # ===================== Internals =====================
    def _load_model_async(self) -> None:
        if YOLO is None:
            messagebox.showerror("错误", f"未安装ultralytics: {_import_error}")
            return
        if not self.weights_path or not self.weights_path.exists():
            messagebox.showerror("错误", "权重文件不存在")
            return
        self.status_text.set("正在加载模型…")

        def _worker() -> None:
            try:
                model = YOLO(str(self.weights_path))
            except Exception as exc:  # pragma: no cover
                self.root.after(0, lambda: messagebox.showerror("错误", f"Loading the model failed: {exc}"))
                self.root.after(0, lambda: self.status_text.set("Loading the model failed"))
                return
            self.model = model
            try:
                self.class_names = getattr(model, "names", None)
            except Exception:
                self.class_names = None
            self.root.after(0, lambda: self.status_text.set("模型已加载，准备检测"))

        threading.Thread(target=_worker, daemon=True).start()

    def _load_cls_model_async(self) -> None:
        if YOLO is None:
            messagebox.showerror("错误", f"未安装ultralytics: {_import_error}")
            return
        if not self.cls_weights_path or not self.cls_weights_path.exists():
            messagebox.showerror("错误", "分类权重文件不存在")
            return
        self.status_text.set("正在加载分类模型…")

        def _worker() -> None:
            try:
                model = YOLO(str(self.cls_weights_path))
            except Exception as exc:
                self.root.after(0, lambda: messagebox.showerror("错误", f"加载分类模型失败: {exc}"))
                self.root.after(0, lambda: self.status_text.set("分类模型加载失败"))
                return
            self.cls_model = model
            try:
                self.cls_class_names = getattr(model, "names", None)
            except Exception:
                self.cls_class_names = None
            self.root.after(0, lambda: self.status_text.set("分类模型已加载"))

        threading.Thread(target=_worker, daemon=True).start()

    def _camera_loop(self) -> None:
        prev_time = time.perf_counter()
        ema_fps = None

        while not self.stop_event.is_set():
            ok, frame = self.capture.read() if self.capture is not None else (False, None)
            if not ok or frame is None:
                time.sleep(0.01)
                continue

            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

            # Inference
            t0 = time.perf_counter()
            try:
                results = self.model.predict(
                    source=frame_rgb, conf=float(self.conf_threshold.get()), verbose=False
                )
            except Exception as exc:  # pragma: no cover
                self.root.after(0, lambda e=exc: self.status_text.set(f"推理错误: {e}"))
                results = []

            annotated_bgr, num_faces, dets = self._draw_detections(frame, results)
            # Tracking disabled

            # FPS and elapsed
            t1 = time.perf_counter()
            dt = t1 - prev_time
            prev_time = t1
            inst_fps = 1.0 / max(dt, 1e-6)
            ema_fps = inst_fps if ema_fps is None else (0.9 * ema_fps + 0.1 * inst_fps)
            elapsed_ms = (t1 - t0) * 1000.0

            self.root.after(0, self.current_fps.set, f"FPS: {ema_fps:.1f}")
            # For live camera, do not treat per-frame time as task time; keep as '-' during session
            self.root.after(0, self.target_info_var.set, f"目标数: {len(dets)}")
            # Camera: cumulative mode (append=True)
            self.root.after(0, self._update_info_table, dets, True)

            # Alarm check
            if self.alarm_enabled.get():
                if num_faces >= int(self.alarm_min_faces.get()):
                    now = time.perf_counter()
                    if now - self._last_alarm_ts > 1.5:
                        self._last_alarm_ts = now
                        if winsound is not None:
                            try:
                                winsound.Beep(1200, 200)
                            except Exception:
                                self.root.bell()
                        else:
                            self.root.bell()

            # Optionally record to history at most once per second
            try:
                if self.record_history_enabled.get():
                    now2 = time.perf_counter()
                    if now2 - self._last_hist_ts > 1.0 and dets:
                        self._last_hist_ts = now2
                        cam_path = f"camera:{int(self.camera_index.get())}"
                        # Limit number of entries per tick to avoid flooding
                        for d in dets[:5]:
                            self._add_history_detection(cam_path, d)
            except Exception:
                pass

            # Display
            disp_rgb = cv2.cvtColor(annotated_bgr, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(disp_rgb)
            # Use fixed maximum dimensions to ensure consistent display size
            max_w = 900
            max_h = 440
            img.thumbnail((max_w, max_h), Image.LANCZOS)
            imgtk = ImageTk.PhotoImage(image=img)

            # Keep reference to avoid GC
            self.canvas.imgtk = imgtk
            self.root.after(0, self.canvas.configure, {"image": imgtk})

            with self.frame_lock:
                self.last_frame_bgr = annotated_bgr

        # loop end cleanup done in on_stop_camera

    def _draw_detections(self, frame_bgr: np.ndarray, results: list):
        annotated = frame_bgr.copy()
        num = 0
        if not results:
            return annotated, num, []
        res = results[0]
        try:
            boxes = res.boxes
        except AttributeError:
            return annotated, num, []
        if boxes is None:
            return annotated, num, []

        xyxy = boxes.xyxy.cpu().numpy() if hasattr(boxes, "xyxy") else None
        conf = boxes.conf.cpu().numpy() if hasattr(boxes, "conf") else None
        cls_ids = boxes.cls.cpu().numpy() if hasattr(boxes, "cls") else None

        if xyxy is None or conf is None:
            return annotated, num, []

        # If segmentation masks are present, use Ultralytics renderer to overlay masks
        try:
            if getattr(res, "masks", None) is not None:
                annotated = res.plot()  # returns BGR with masks/boxes/labels
        except Exception:
            annotated = frame_bgr.copy()

        detections = []
        for idx, ((x1, y1, x2, y2), c) in enumerate(zip(xyxy.astype(int), conf)):
            # Determine class name first
            cls_val = None
            cls_name = None
            if cls_ids is not None and idx < len(cls_ids):
                try:
                    cls_val = int(cls_ids[idx])
                except Exception:
                    cls_val = None
            if cls_val is not None:
                try:
                    names = getattr(self, "class_names", None)
                    if isinstance(names, (list, tuple, dict)):
                        if isinstance(names, dict):
                            cls_name = names.get(cls_val, str(cls_val))
                        else:
                            if 0 <= cls_val < len(names):
                                cls_name = names[cls_val]
                except Exception:
                    cls_name = None

            # If no masks, draw boxes/labels ourselves with class name
            if getattr(res, "masks", None) is None:
                color = (0, 200, 255)
                cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 2)
                # Run classification on face crop if classification model is available
                emotion_txt = None
                emotion_cls_name = None
                emotion_conf = None
                if getattr(self, "cls_model", None) is not None:
                    try:
                        x1c = max(0, x1); y1c = max(0, y1); x2c = min(frame_bgr.shape[1]-1, x2); y2c = min(frame_bgr.shape[0]-1, y2)
                        if x2c > x1c and y2c > y1c:
                            face_crop = frame_bgr[y1c:y2c, x1c:x2c]
                            crop_rgb = cv2.cvtColor(face_crop, cv2.COLOR_BGR2RGB)
                            cls_res = self.cls_model.predict(source=crop_rgb, verbose=False)
                            if cls_res and hasattr(cls_res[0], "probs") and cls_res[0].probs is not None:
                                probs = cls_res[0].probs
                                top_i = int(getattr(probs, "top1", None) if getattr(probs, "top1", None) is not None else int(np.argmax(probs.data.cpu().numpy())))
                                top_p = float(getattr(probs, "top1conf", None) if getattr(probs, "top1conf", None) is not None else float(np.max(probs.data.cpu().numpy())))
                                names_em = getattr(self, "cls_class_names", None)
                                if names_em is None:
                                    try:
                                        names_em = getattr(cls_res[0], "names", None)
                                    except Exception:
                                        names_em = None
                                if isinstance(names_em, dict):
                                    emotion_cls_name = names_em.get(top_i, str(top_i))
                                    emotion_txt = f"{emotion_cls_name} {top_p:.2f}"
                                elif isinstance(names_em, (list, tuple)) and 0 <= top_i < len(names_em):
                                    emotion_cls_name = names_em[top_i]
                                    emotion_txt = f"{emotion_cls_name} {top_p:.2f}"
                                else:
                                    emotion_cls_name = str(top_i)
                                    emotion_txt = f"{emotion_cls_name} {top_p:.2f}"
                                emotion_conf = top_p
                    except Exception:
                        emotion_txt = None

                base_txt = f"{cls_name} {c:.2f}" if cls_name is not None else f"{c:.2f}"
                label_txt = (emotion_txt or base_txt)
                (tw, th), _ = cv2.getTextSize(label_txt, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                cv2.rectangle(annotated, (x1, max(0, y1 - th - 6)), (x1 + tw + 4, y1), color, -1)
                cv2.putText(annotated, label_txt, (x1 + 2, max(0, y1 - 4)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 2)

            num += 1
            det = {"box": [int(x1), int(y1), int(x2), int(y2)], "conf": float(c)}
            if cls_val is not None:
                det["cls"] = cls_val
            if cls_name is not None:
                det["cls_name"] = cls_name
            if 'emotion_txt' in locals() and emotion_txt:
                det["emotion"] = emotion_txt
            if 'emotion_cls_name' in locals() and emotion_cls_name is not None:
                det["emotion_cls"] = emotion_cls_name
            if 'emotion_conf' in locals() and emotion_conf is not None:
                det["emotion_conf"] = float(emotion_conf)
            detections.append(det)

        return annotated, num, detections

    # ===================== Image/Video/Batch/Export =====================
    def _predict_on_image(self, bgr, update_table=True, append=False):
        """
        Predict on a single image.
        Args:
            bgr: Input image in BGR format
            update_table: Whether to update the detection info table
            append: If True, append to table; if False, replace table (only used if update_table=True)
        """
        t0 = time.perf_counter()
        frame_rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        results = self.model.predict(source=frame_rgb, conf=float(self.conf_threshold.get()), verbose=False)
        annotated, num, dets = self._draw_detections(bgr, results)
        if update_table:
            self._update_info_table(dets, append=append)
        elapsed_ms = (time.perf_counter() - t0) * 1000.0
        try:
            self.time_info_var.set(f"用时: {elapsed_ms:.1f} ms")
        except Exception:
            pass
        return annotated, dets

    def _add_history_detection(self, path: Union[str, Path], det: Dict) -> None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Prefer emotion fields if available
        cls_display = det.get("emotion_cls") if det.get("emotion_cls") is not None else (
            det.get("cls_name") if det.get("cls_name") is not None else det.get("cls", "")
        )
        conf = det.get("emotion_conf", det.get("conf", 0.0))
        box = det.get("box", [0, 0, 0, 0])
        bbox_txt = f"[{box[0]}, {box[1]}, {box[2]}, {box[3]}]"
        entry = {
            "idx": self.history_counter,
            "path": str(path),
            "emotion": cls_display,
            "emotion_conf": float(conf),
            "bbox": box,
            "time": ts,
        }
        self.history.append(entry)
        self.history_tv.insert("", tk.END, values=(entry["idx"], entry["path"], cls_display, f"{conf:.2f}", bbox_txt, ts))
        self.history_counter += 1

    def on_detect_image(self) -> None:
        if self.model is None:
            messagebox.showwarning("提示", "请先选择并加载检测模型权重！")
            return
        paths = filedialog.askopenfilenames(title="选择图片", filetypes=[("Images", "*.jpg;*.jpeg;*.png;*.bmp"), ("All Files", "*.*")])
        if not paths:
            return
        # If multiple selected, treat as batch
        if len(paths) > 1:
            self._start_batch_with_progress(list(paths))
            return
        img_path = paths[0]
        bgr = cv2.imread(img_path)
        if bgr is None:
            messagebox.showerror("错误", f"无法读取图片: {img_path}")
            return
        # Reset detection counter and clear table for new image detection (single image: replace mode)
        self.detection_counter = 1
        self.dets_tv.delete(*self.dets_tv.get_children())
        annotated, dets = self._predict_on_image(bgr)

        # 统计检测目标数量
        num_dets = len(dets)
        for d in dets:
            self._add_history_detection(img_path, d)
        # 更新检测目标数量显示
        self.target_info_var.set(f"目标数: {num_dets}")
        self._show_image_on_canvas(annotated)
        # Store for manual save
        self.last_annotated_image = annotated
        self.last_batch_dir = None
        self.save_result_btn.config(state=tk.NORMAL)

    def on_capture_photo_detect(self) -> None:
        if self.model is None:
            messagebox.showwarning("提示", "请先选择并加载模型权重！")
            return
        # Open a preview window to allow pose adjustment
        if getattr(self, "photo_preview_win", None) is not None:
            try:
                self.photo_preview_win.lift()
                return
            except Exception:
                pass
        self.photo_preview_win = tk.Toplevel(self.root)
        self.photo_preview_win.title("拍照预览")
        self.photo_preview_win.geometry("640x520")
        self.photo_preview_label = tk.Label(self.photo_preview_win, bg="#000")
        self.photo_preview_label.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        btns = ttk.Frame(self.photo_preview_win)
        btns.pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Button(btns, text="拍照", command=self._photo_take_snapshot).pack(side=tk.LEFT, padx=8, pady=8)
        ttk.Button(btns, text="关闭", command=self._photo_close_preview).pack(side=tk.RIGHT, padx=8, pady=8)
        # Setup camera
        index = int(self.camera_index.get())
        self.photo_preview_cap = cv2.VideoCapture(index, cv2.CAP_DSHOW)
        if not self.photo_preview_cap.isOpened():
            messagebox.showerror("错误", f"无法打开摄像头 {index}")
            try:
                self.photo_preview_win.destroy()
            except Exception:
                pass
            self.photo_preview_win = None
            return
        self.photo_last_frame = None
        self._photo_preview_loop()
        self.photo_preview_win.protocol("WM_DELETE_WINDOW", self._photo_close_preview)

    def _photo_preview_loop(self) -> None:
        if getattr(self, "photo_preview_win", None) is None:
            return
        ok, frame = (False, None)
        try:
            if self.photo_preview_cap is not None:
                ok, frame = self.photo_preview_cap.read()
        except Exception:
            ok = False
        if ok and frame is not None:
            self.photo_last_frame = frame
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(rgb)
            max_w = self.photo_preview_label.winfo_width() or 640
            max_h = self.photo_preview_label.winfo_height() or 480
            img.thumbnail((max_w, max_h), Image.LANCZOS)
            imgtk = ImageTk.PhotoImage(image=img)
            self.photo_preview_label.imgtk = imgtk
            self.photo_preview_label.configure(image=imgtk)
        # schedule next update
        if getattr(self, "photo_preview_win", None) is not None:
            self.root.after(30, self._photo_preview_loop)

    def _photo_take_snapshot(self) -> None:
        frame = getattr(self, "photo_last_frame", None)
        if frame is None:
            messagebox.showinfo("提示", "未捕获到画面，请稍后再试")
            return
        # Reset detection counter and clear table for new photo detection (single photo: replace mode)
        self.detection_counter = 1
        self.dets_tv.delete(*self.dets_tv.get_children())
        annotated, dets = self._predict_on_image(frame)
        cam_path = f"camera:{int(self.camera_index.get())}"
        for d in dets:
            self._add_history_detection(cam_path, d)
        self._show_image_on_canvas(annotated)
        # Store for manual save
        self.last_annotated_image = annotated
        self.last_batch_dir = None
        self.save_result_btn.config(state=tk.NORMAL)

    def _photo_close_preview(self) -> None:
        try:
            if getattr(self, "photo_preview_cap", None) is not None:
                self.photo_preview_cap.release()
        except Exception:
            pass
        self.photo_preview_cap = None
        try:
            if getattr(self, "photo_preview_win", None) is not None:
                self.photo_preview_win.destroy()
        except Exception:
            pass
        self.photo_preview_win = None

    def on_select_video_detect(self) -> None:
        if self.model is None:
            messagebox.showwarning("提示", "请先选择并加载模型权重！")
            return
        path = filedialog.askopenfilename(title="选择视频文件", filetypes=[("Videos", "*.mp4;*.avi;*.mov;*.mkv"), ("All Files", "*.*")])
        if not path:
            return
        # Disable controls during processing
        self._set_file_controls_state(tk.DISABLED)
        self.status_text.set("视频检测中…")

        def _video_preview_worker():
            t_task0 = time.perf_counter()
            cap = None
            try:
                cap = cv2.VideoCapture(path)
                if not cap.isOpened():
                    self.root.after(0, lambda: messagebox.showerror("错误", f"无法打开视频: {path}"))
                    return
                total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
                processed = 0
                faces_sum = 0
                # Reset detection counter and clear table for new video (video: cumulative mode)
                self.detection_counter = 1
                self.root.after(0, lambda: self.dets_tv.delete(*self.dets_tv.get_children()))
                # Tracking disabled for video preview
                prev_time = time.perf_counter()
                ema_fps = None
                while True:
                    ok, frame = cap.read()
                    if not ok:
                        break
                    t0 = time.perf_counter()
                    results = self.model.predict(source=cv2.cvtColor(frame, cv2.COLOR_BGR2RGB), conf=float(self.conf_threshold.get()), verbose=False)
                    annotated, num, dets = self._draw_detections(frame, results)
                    t1 = time.perf_counter()
                    # Tracking disabled
                    faces_sum += num
                    processed += 1
                    # simple FPS estimate and elapsed per frame
                    dt = t1 - prev_time
                    prev_time = t1
                    inst_fps = 1.0 / max(dt, 1e-6)
                    ema_fps = inst_fps if ema_fps is None else (0.9 * ema_fps + 0.1 * inst_fps)
                    elapsed_ms = (t1 - t0) * 1000.0

                    # Add frame number to each detection for history
                    frame_path = f"{path} (帧 {processed})"
                    # Prepare detections with frame number
                    dets_with_frame = []
                    for d in dets:
                        d_with_frame = d.copy()
                        d_with_frame["frame_num"] = processed
                        dets_with_frame.append(d_with_frame)

                    # Batch update: add all detections from this frame to history and table
                    if dets_with_frame:
                        def _batch_add_frame_dets(frame_p, detections):
                            for d in detections:
                                self._add_history_detection(frame_p, d)
                                self._update_info_table([d], True)
                        self.root.after(0, _batch_add_frame_dets, frame_path, dets_with_frame)

                    if processed % 5 == 0:
                        # only update image widget to avoid touching other widgets
                        self.root.after(0, self._show_image_on_canvas, annotated)
                    if processed % 20 == 0:
                        # batch update of text-only labels to minimize redraws
                        def _batch_text_update(p, tot, fpsv, tgt):
                            try:
                                self.status_text.set(f"视频检测中… {p}/{tot} 帧")
                                if fpsv is not None:
                                    self.current_fps.set(f"FPS: {fpsv:.1f}")
                                self.target_info_var.set(f"目标数: {tgt}")
                            except Exception:
                                pass
                        self.root.after(0, _batch_text_update, processed, total_frames, ema_fps, len(dets))
                # End of video task: update total elapsed time
                t_task1 = time.perf_counter()
                total_ms = (t_task1 - t_task0) * 1000.0
                self.root.after(0, self.time_info_var.set, f"用时: {total_ms:.1f} ms")
                # Arm save state
                self.last_annotated_image = None
                self.last_batch_dir = None
                self.last_video_path = path
                self.root.after(0, self.save_result_btn.config, {"state": tk.NORMAL})
                self.root.after(0, self.status_text.set, "视频检测完成，可点击'保存检测结果'导出整段视频")
            finally:
                try:
                    if cap is not None:
                        cap.release()
                except Exception:
                    pass
                self.root.after(0, self._set_file_controls_state, tk.NORMAL)

        threading.Thread(target=_video_preview_worker, daemon=True).start()

    def _detect_batch_images(self, paths, on_progress: Optional[Callable[[int, int], None]] = None, on_done: Optional[Callable[[Path, float], None]] = None) -> None:
        # Process but do not auto-save; keep to temp in memory and allow user to save
        t_task0 = time.perf_counter()
        saved_dir = self.output_dir / f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        saved_dir.mkdir(parents=True, exist_ok=True)
        total_faces = 0
        total = len(paths)
        processed = 0
        # Reset detection counter and clear table for batch detection (batch: cumulative mode)
        self.detection_counter = 1
        self.root.after(0, lambda: self.dets_tv.delete(*self.dets_tv.get_children()))
        for p in paths:
            bgr = cv2.imread(p)
            if bgr is None:
                processed += 1
                if on_progress:
                    try:
                        on_progress(processed, total)
                    except Exception:
                        pass
                continue
            # Batch: don't update table in _predict_on_image, we'll do it manually in cumulative mode
            annotated, dets = self._predict_on_image(bgr, update_table=False)
            total_faces += len(dets)
            out = saved_dir / (Path(p).stem + "_det.jpg")
            cv2.imwrite(str(out), annotated)
            for d in dets:
                self._add_history_detection(p, d)
            # Batch: cumulative mode (append to table)
            if dets:
                self.root.after(0, self._update_info_table, dets, True)
            processed += 1
            if on_progress:
                try:
                    on_progress(processed, total)
                except Exception:
                    pass
        total_ms = (time.perf_counter() - t_task0) * 1000.0
        if on_done is not None:
            try:
                on_done(saved_dir, total_ms)
            except Exception:
                pass
        else:
            self.status_text.set(f"批量检测完成，可点击'保存检测结果'选择导出位置")
            try:
                self.time_info_var.set(f"用时: {total_ms:.1f} ms")
            except Exception:
                pass
            self.last_annotated_image = None
            self.last_batch_dir = saved_dir
            self.save_result_btn.config(state=tk.NORMAL)

    def on_batch_images_detect(self) -> None:
        if self.model is None:
            messagebox.showwarning("提示", "请先选择并加载检测模型权重！")
            return
        paths = filedialog.askopenfilenames(title="选择多张图片", filetypes=[("Images", "*.jpg;*.jpeg;*.png;*.bmp"), ("All Files", "*.*")])
        if not paths:
            return
        self._start_batch_with_progress(list(paths))

    def _start_batch_with_progress(self, paths: List[str]) -> None:
        # Modal progress dialog
        prog = tk.Toplevel(self.root)
        prog.title("批量检测进度")
        prog.geometry("360x120")
        prog.transient(self.root)
        prog.grab_set()
        ttk.Label(prog, text=f"共 {len(paths)} 张图片").pack(side=tk.TOP, pady=(12, 6))
        pb = ttk.Progressbar(prog, orient=tk.HORIZONTAL, mode="determinate", maximum=len(paths))
        pb.pack(fill=tk.X, padx=16, pady=(0, 8))
        msg_var = tk.StringVar(value="准备开始…")
        ttk.Label(prog, textvariable=msg_var).pack(side=tk.TOP)

        self._set_file_controls_state(tk.DISABLED)

        def on_progress(done: int, total: int):
            def _ui():
                try:
                    pb['value'] = done
                    msg_var.set(f"已完成 {done}/{total}")
                except Exception:
                    pass
            self.root.after(0, _ui)

        def on_done(saved_dir: Path, total_ms: float):
            def _ui_done():
                try:
                    self.status_text.set("批量检测完成，可点击'保存检测结果'选择导出位置")
                    self.time_info_var.set(f"用时: {total_ms:.1f} ms")
                    self.last_annotated_image = None
                    self.last_batch_dir = saved_dir
                    self.save_result_btn.config(state=tk.NORMAL)
                except Exception:
                    pass
                finally:
                    try:
                        prog.destroy()
                    except Exception:
                        pass
                    self._set_file_controls_state(tk.NORMAL)
                    messagebox.showinfo("提示", "批量检测完成！")
            self.root.after(0, _ui_done)

        def _worker():
            try:
                self._detect_batch_images(paths, on_progress=on_progress, on_done=on_done)
            except Exception as exc:
                def _ui_err(e=exc):
                    try:
                        prog.destroy()
                    except Exception:
                        pass
                    self._set_file_controls_state(tk.NORMAL)
                    messagebox.showerror("错误", f"批量检测失败: {e}")
                self.root.after(0, _ui_err)

        threading.Thread(target=_worker, daemon=True).start()

    def on_export_results(self) -> None:
        if not self.history:
            messagebox.showinfo("提示", "暂无历史记录可导出")
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV 文件", "*.csv")], initialfile="face_results.csv")
        if not path:
            return
        import csv
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["idx", "path", "emotion", "emotion_conf", "bbox", "time"])
            w.writeheader()
            for row in self.history:
                w.writerow(row)
        self.status_text.set(f"已导出CSV: {path}")

    def on_export_detection_info(self) -> None:
        """Export detection info table data to CSV file."""
        # Get all items from the detection info table
        items = self.dets_tv.get_children()
        if not items:
            messagebox.showinfo("提示", "检测信息表格为空，无可导出数据")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv")],
            initialfile="detection_info.csv"
        )
        if not path:
            return

        import csv
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                # Table columns: idx, emotion, emotion_conf, x1, y1, x2, y2
                w = csv.writer(f)
                # Write header
                w.writerow(["序号", "表情", "表情置信度", "x1", "y1", "x2", "y2"])
                # Write data rows
                for item in items:
                    values = self.dets_tv.item(item, "values")
                    if values:
                        w.writerow(values)
            self.status_text.set(f"已导出检测信息CSV: {path}")
            messagebox.showinfo("提示", f"检测信息已成功导出到: {path}")
        except Exception as exc:
            messagebox.showerror("错误", f"导出失败: {exc}")

    def _show_image_on_canvas(self, bgr: np.ndarray) -> None:
        disp_rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(disp_rgb)
        # Use fixed maximum dimensions to ensure consistent display size
        # Canvas is set to 960x540, but we use pixel dimensions for image scaling
        max_w = 900
        max_h = 440
        img.thumbnail((max_w, max_h), Image.LANCZOS)
        imgtk = ImageTk.PhotoImage(image=img)
        self.canvas.imgtk = imgtk
        self.canvas.configure(image=imgtk)

    def _set_file_controls_state(self, state) -> None:
        try:
            for btn in self._file_buttons:
                btn.config(state=state)
        except Exception:
            pass

    def _update_info_table(self, dets, append=False) -> None:
        """
        Update the detection info table.
        Args:
            dets: List of detection dictionaries
            append: If True, append to existing rows; if False, replace all rows
        """
        try:
            if not append:
                self.dets_tv.delete(*self.dets_tv.get_children())
            for d in dets or []:
                box = d.get("box", [0, 0, 0, 0])
                emo_conf = d.get("emotion_conf", None)
                conf_display = f"{emo_conf:.2f}" if isinstance(emo_conf, (int, float)) else "-"
                emo_name = d.get("emotion_cls")
                cls_display = emo_name if emo_name is not None else "-"
                # Use detection_counter for row index
                idx = self.detection_counter
                self.dets_tv.insert("", tk.END, values=(idx, cls_display, conf_display, box[0], box[1], box[2], box[3]))
                # Always increment counter for next detection
                self.detection_counter += 1
        except Exception:
            pass

    def _update_weights_display(self) -> None:
        # Shorten filename and attach tooltip with full path; allow wrap to next line if needed
        try:
            if not getattr(self, "weights_path", None):
                self.weights_label_var.set("Weight: Not selected")
                self._set_tooltip(self.weights_label, "Not chosen")
                return
            name = self.weights_path.name
            max_len = 22
            short = name if len(name) <= max_len else (name[:max_len - 3] + "...")
            self.weights_label_var.set(f"Weight: {short}")
            self._set_tooltip(self.weights_label, str(self.weights_path))
        except Exception:
            self.weights_label_var.set("Weight: Not selected")
            self._set_tooltip(self.weights_label, "Not chosen")

    def _update_cls_weights_display(self) -> None:
        try:
            if not getattr(self, "cls_weights_path", None):
                self.cls_weights_label_var.set("Classification weight: Not selected")
                self._set_tooltip(self.cls_weights_label, "Not chosen")
                return
            name = self.cls_weights_path.name
            max_len = 22
            short = name if len(name) <= max_len else (name[:max_len - 3] + "...")
            self.cls_weights_label_var.set(f"Classification weight: {short}")
            self._set_tooltip(self.cls_weights_label, str(self.cls_weights_path))
        except Exception:
            self.cls_weights_label_var.set("Classification weight: Not selected")
            self._set_tooltip(self.cls_weights_label, "Not chosen")

    def _set_tooltip(self, widget, text: str) -> None:
        # Lightweight tooltip for ttk widgets
        try:
            if hasattr(widget, "_tooltip") and widget._tooltip is not None:
                try:
                    widget._tooltip.destroy()
                except Exception:
                    pass
                widget._tooltip = None

            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.withdraw()
            tip.attributes("-topmost", True)
            lbl = ttk.Label(tip, text=text, background="#FFFFE0", relief=tk.SOLID, borderwidth=1)
            lbl.pack(ipadx=6, ipady=3)
            widget._tooltip = tip

            def show_tip(event=None):
                try:
                    x = widget.winfo_rootx() + 10
                    y = widget.winfo_rooty() + widget.winfo_height() + 6
                    tip.geometry(f"+{x}+{y}")
                    tip.deiconify()
                except Exception:
                    pass

            def hide_tip(event=None):
                try:
                    tip.withdraw()
                except Exception:
                    pass

            widget.bind("<Enter>", show_tip)
            widget.bind("<Leave>", hide_tip)
        except Exception:
            pass

    def on_save_detection_result(self) -> None:
        # Save the last annotated single image or copy the last batch dir
        if self.last_annotated_image is not None:
            path = filedialog.asksaveasfilename(defaultextension=".jpg", filetypes=[("JPEG", "*.jpg"), ("PNG", "*.png")], initialfile="detected.jpg")
            if not path:
                return
            try:
                cv2.imwrite(str(path), self.last_annotated_image)
                self.status_text.set(f"已保存: {path}")
            except Exception as exc:
                messagebox.showerror("错误", f"保存失败: {exc}")
            return
        # Save annotated full video by reprocessing last video path (run in background to avoid freezing UI)
        if getattr(self, "last_video_path", None):
            src = self.last_video_path
            default_name = f"video_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp4"
            dest = filedialog.asksaveasfilename(
                title="保存标注视频为…",
                defaultextension=".mp4",
                initialfile=default_name,
                filetypes=[("MP4", "*.mp4"), ("AVI", "*.avi"), ("All Files", "*.*")]
            )
            if not dest:
                return
            self._set_file_controls_state(tk.DISABLED)

            def _video_save_worker():
                cap2 = None
                writer = None
                try:
                    cap2 = cv2.VideoCapture(src)
                    if not cap2.isOpened():
                        self.root.after(0, lambda: messagebox.showerror("错误", f"无法打开视频: {src}"))
                        return
                    w = int(cap2.get(cv2.CAP_PROP_FRAME_WIDTH))
                    h = int(cap2.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    fps = cap2.get(cv2.CAP_PROP_FPS) or 25.0
                    fourcc = cv2.VideoWriter_fourcc(*("mp4v" if dest.lower().endswith(".mp4") else "XVID"))
                    writer = cv2.VideoWriter(str(dest), fourcc, fps, (w, h))
                    if not writer.isOpened():
                        self.root.after(0, lambda: messagebox.showerror("错误", "无法创建视频写入器"))
                        return
                    total_frames = int(cap2.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
                    processed = 0
                    # Tracking disabled for video saving
                    while True:
                        ok, frame = cap2.read()
                        if not ok:
                            break
                        results = self.model.predict(source=cv2.cvtColor(frame, cv2.COLOR_BGR2RGB), conf=float(self.conf_threshold.get()), verbose=False)
                        annotated, num, dets = self._draw_detections(frame, results)
                        # Tracking disabled
                        writer.write(annotated)
                        processed += 1
                        if processed % 20 == 0:
                            self.root.after(0, self.status_text.set, f"保存视频中… {processed}/{total_frames} 帧")
                    self.root.after(0, self.status_text.set, f"视频已保存: {dest}")
                finally:
                    try:
                        if writer is not None:
                            writer.release()
                    except Exception:
                        pass
                    try:
                        if cap2 is not None:
                            cap2.release()
                    except Exception:
                        pass
                    self.root.after(0, self._set_file_controls_state, tk.NORMAL)

            threading.Thread(target=_video_save_worker, daemon=True).start()
            return
        if self.last_batch_dir is not None and self.last_batch_dir.exists():
            dest = filedialog.askdirectory(title="选择导出文件夹")
            if not dest:
                return
            dest_path = Path(dest) / self.last_batch_dir.name
            try:
                if dest_path.exists():
                    messagebox.showerror("错误", f"目标已存在: {dest_path}")
                    return
                shutil.copytree(self.last_batch_dir, dest_path)
                self.status_text.set(f"已导出到: {dest_path}")
            except Exception as exc:
                messagebox.showerror("错误", f"导出失败: {exc}")
            return
        messagebox.showinfo("提示", "暂无可保存的检测结果")


class EmotionAnalysisWindow:
    def __init__(self, parent_app: FaceDetectionApp) -> None:
        self.parent_app = parent_app
        self.root = tk.Toplevel(parent_app.root)
        self.root.title("表情分析")
        self.root.geometry("1200x720")
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Menu to switch back
        try:
            menubar = tk.Menu(self.root)
            func_menu = tk.Menu(menubar, tearoff=0)
            func_menu.add_command(label="face detection", command=self._back_to_face)
            menubar.add_cascade(label="function", menu=func_menu)
            # Help menu for emotion window
            help_menu = tk.Menu(menubar, tearoff=0)
            help_menu.add_command(label="about", command=lambda: messagebox.showinfo("关于", "Expression Analysis Module"))
            help_menu.add_separator()
            help_menu.add_command(label="esc", command=self.parent_app.root.destroy)
            menubar.add_cascade(label="help", menu=help_menu)
            self.root.config(menu=menubar)
        except Exception:
            pass

        # Top controls
        controls = ttk.Frame(self.root)
        controls.pack(side=tk.TOP, fill=tk.X, padx=8, pady=8)

        ttk.Label(controls, text="Name of the large model").pack(side=tk.LEFT)
        self.model_name_var = tk.StringVar(value="Qwen/Qwen2.5-VL-72B-Instruct")
        ttk.Entry(controls, textvariable=self.model_name_var, width=28).pack(side=tk.LEFT, padx=(6, 16))

        # 使用支持自定义颜色的主题
        style = ttk.Style()
        style.configure("Purple.TButton", background="#9C27B0", foreground="black")   # 紫色
        style.configure("Teal.TButton",   background="#009688", foreground="black")   # 蓝绿色
        style.configure("Gray.TButton",   background="#9E9E9E", foreground="black")   # 中灰
        ttk.Label(controls, text="image file").pack(side=tk.LEFT)
        self.image_path_var = tk.StringVar(value="")
        ttk.Entry(controls, textvariable=self.image_path_var, width=50).pack(side=tk.LEFT, padx=6)
        ttk.Button(controls, text="浏览...", style="Purple.TButton", command=self._on_browse_image).pack(side=tk.LEFT, padx=(6, 16))

        self.analyze_btn = ttk.Button(controls, text="Expression analysis", style="Teal.TButton", command=self._on_analyze)
        self.analyze_btn.pack(side=tk.LEFT)
        ttk.Button(controls, text="Export the analysis results", style="Gray.TButton", command=self._on_export_results).pack(side=tk.LEFT, padx=(12, 0))

        # Body: left image, right results
        body = ttk.Frame(self.root)
        body.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        self.canvas = tk.Label(body, bg="#111111")
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        right = ttk.LabelFrame(body, text="Analysis results")
        right.pack(side=tk.RIGHT, fill=tk.Y)
        # Use tree + headings so we can display thumbnail in the tree column (#0)
        # Ensure rows are tall enough for thumbnails to avoid overlap
        try:
            _style = ttk.Style()
            # Row height will be set after defining thumbnail size below
            tree_style = "Thumb.Treeview"
        except Exception:
            tree_style = None
        # Thumbnail size for right panel list
        self.thumb_size = 96
        try:
            if tree_style:
                _style.configure(tree_style, rowheight=self.thumb_size + 8)
        except Exception:
            pass
        self.results_tv = ttk.Treeview(right, columns=("idx", "bbox", "emotion"), show="tree headings", height=8, style=(tree_style or ""))
        self.results_tv.heading("#0", text="Face")
        self.results_tv.column("#0", width=self.thumb_size + 8, anchor=tk.W)
        for col, txt, w in (("idx", "Serial Number", 60), ("bbox", "Face frame", 260), ("emotion", "Emotion", 360)):
            self.results_tv.heading(col, text=txt)
            self.results_tv.column(col, width=w, anchor=tk.W)
        self.results_tv.pack(side=tk.LEFT, fill=tk.Y, padx=(8, 0), pady=(8, 8))
        sb = ttk.Scrollbar(right, orient=tk.VERTICAL, command=self.results_tv.yview)
        self.results_tv.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y, pady=(8, 8))

        # State
        self._current_bgr = None  # type: Optional[np.ndarray]
        self._last_annotated = None  # type: Optional[np.ndarray]
        self._thumb_refs = []  # keep thumbnail PhotoImage refs
        self._result_items = []  # rows for export: list of dicts with idx, bbox, emotion, thumb

        # History section at bottom
        hist_frame = ttk.LabelFrame(self.root, text="Historical record")
        # Ensure history bar reserves space by packing before the large body frame
        hist_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=8, pady=(0, 8), before=body)
        self.history_tv2 = ttk.Treeview(hist_frame, columns=("idx", "model", "path", "total_ms", "finished_at"), show="headings", height=6)
        for col, txt, w in (
            ("idx", "Serial Number", 60),
            ("model", "Name of the large model", 220),
            ("path", "Image file path", 520),
            ("total_ms", "Time taken (ms)", 120),
            ("finished_at", "Analysis time", 180),
        ):
            self.history_tv2.heading(col, text=txt)
            self.history_tv2.column(col, width=w, anchor=tk.W)
        self.history_tv2.pack(side=tk.LEFT, fill=tk.X, expand=True)
        hist_sb = ttk.Scrollbar(hist_frame, orient=tk.VERTICAL, command=self.history_tv2.yview)
        self.history_tv2.configure(yscrollcommand=hist_sb.set)
        hist_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._hist_counter = 1

    def _on_close(self) -> None:
        try:
            self.root.destroy()
        finally:
            try:
                self.parent_app.root.deiconify()
                self.parent_app.root.lift()
            except Exception:
                pass

    def _back_to_face(self) -> None:
        try:
            self.root.destroy()
        finally:
            self.parent_app._switch_to_face_detection()

    def _on_browse_image(self) -> None:
        path = filedialog.askopenfilename(title="选择图片", filetypes=[("Images", "*.jpg;*.jpeg;*.png;*.bmp"), ("All Files", "*.*")])
        if not path:
            return
        self.image_path_var.set(path)
        bgr = cv2.imread(path)
        if bgr is None:
            messagebox.showerror("错误", f"无法读取图片: {path}")
            return
        self._current_bgr = bgr
        self._show_image(bgr)

    def _show_image(self, bgr: np.ndarray) -> None:
        try:
            rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(rgb)
            max_w = self.canvas.winfo_width() or 960
            max_h = self.canvas.winfo_height() or 540
            img.thumbnail((max_w, max_h), Image.LANCZOS)
            imgtk = ImageTk.PhotoImage(image=img)
            self.canvas.imgtk = imgtk
            self.canvas.configure(image=imgtk)
        except Exception:
            pass

    def _on_analyze(self) -> None:
        if self.parent_app.model is None:
            messagebox.showwarning("提示", "请先在主界面选择并加载模型权重！")
            return
        if self._current_bgr is None:
            path = self.image_path_var.get().strip()
            if not path:
                messagebox.showinfo("提示", "请先选择图片")
                return
            bgr = cv2.imread(path)
            if bgr is None:
                messagebox.showerror("错误", f"无法读取图片: {path}")
                return
            self._current_bgr = bgr
            self._show_image(bgr)

        self.analyze_btn.config(state=tk.DISABLED)
        threading.Thread(target=self._analyze_worker, daemon=True).start()

    def _on_export_results(self) -> None:
        if not getattr(self, "_result_items", None):
            messagebox.showinfo("提示", "暂无可导出的分析结果")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
        except Exception as exc:
            messagebox.showerror("错误", f"需要安装 openpyxl 才能导出: {exc}")
            return
        path = filedialog.asksaveasfilename(
            title="导出为 Excel", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="emotion_results.xlsx"
        )
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        # headers
        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=2, value="人脸缩略图")
        ws.cell(row=1, column=3, value="人脸框")
        ws.cell(row=1, column=4, value="情绪")
        # autosize column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 48

        row = 2
        for item in self._result_items:
            ws.cell(row=row, column=1, value=item.get("idx"))
            ws.cell(row=row, column=3, value=item.get("bbox"))
            ws.cell(row=row, column=4, value=item.get("emotion"))
            # Embed real thumbnail file if available
            thumb_p = item.get("thumb_path")
            if thumb_p and os.path.exists(thumb_p):
                try:
                    xlimg = XLImage(thumb_p)
                    cell_ref = f"B{row}"
                    ws.add_image(xlimg, cell_ref)
                    ws.row_dimensions[row].height = 80
                except Exception:
                    pass
            row += 1
        try:
            wb.save(path)
        except Exception as exc:
            messagebox.showerror("错误", f"导出失败: {exc}")
            return
        messagebox.showinfo("提示", f"已导出: {path}")

    def _encode_bgr_to_data_url(self, bgr: np.ndarray) -> str:
        rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(rgb)
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/png;base64,{b64}"

    def _yolo_detect_faces(self, bgr: np.ndarray) -> List[Dict]:
        model = self.parent_app.model
        if model is None:
            return []
        try:
            res = model.predict(source=cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB), conf=float(self.parent_app.conf_threshold.get()), verbose=False)
        except Exception:
            return []
        if not res:
            return []
        r0 = res[0]
        try:
            boxes = r0.boxes
        except AttributeError:
            return []
        if boxes is None:
            return []
        xyxy = boxes.xyxy.cpu().numpy() if hasattr(boxes, "xyxy") else None
        conf = boxes.conf.cpu().numpy() if hasattr(boxes, "conf") else None
        cls_ids = boxes.cls.cpu().numpy() if hasattr(boxes, "cls") else None
        if xyxy is None or conf is None:
            return []
        # Prefer filtering by class name that contains 'face' if names are available
        names = getattr(self.parent_app, "class_names", None)
        def is_face_class(idx: Optional[int]) -> bool:
            try:
                if names is None or idx is None:
                    return True  # fallback: keep all
                if isinstance(names, dict):
                    name = str(names.get(int(idx), ""))
                elif isinstance(names, (list, tuple)):
                    if 0 <= int(idx) < len(names):
                        name = str(names[int(idx)])
                    else:
                        name = ""
                else:
                    return True
                name_l = name.lower()
                return ("face" in name_l) or ("head" in name_l)  # common label variants
            except Exception:
                return True

        dets_all = []
        for i, ((x1, y1, x2, y2), c) in enumerate(zip(xyxy.astype(int), conf)):
            cls_val = None
            if cls_ids is not None and i < len(cls_ids):
                try:
                    cls_val = int(cls_ids[i])
                except Exception:
                    cls_val = None
            det = {"box": [int(x1), int(y1), int(x2), int(y2)], "conf": float(c)}
            if cls_val is not None:
                det["cls"] = cls_val
            dets_all.append(det)

        # First try face-filtered boxes
        dets_face = [d for idx, d in enumerate(dets_all) if is_face_class(int(cls_ids[idx]) if (cls_ids is not None and idx < len(cls_ids)) else None)]
        if dets_face:
            return dets_face
        # If no face-like class but there are detections, use all detections as faces
        if dets_all:
            return dets_all
        return []

    def _call_qwen(self, model_name: str, data_url: str) -> str:
        # Returns raw content string
        try:
            from openai import OpenAI  # type: ignore
        except Exception as exc:
            raise RuntimeError(f"OpenAI 客户端不可用: {exc}")
        api_key = os.getenv("SILICONFLOW_API_KEY")
        if not api_key:
            raise RuntimeError("未设置 SILICONFLOW_API_KEY 环境变量。请在系统环境变量中配置 SILICONFLOW_API_KEY 并重启程序。")
        base_url = os.getenv("SILICONFLOW_BASE_URL", "https://api.siliconflow.cn/v1")
        client = OpenAI(api_key=api_key, base_url=base_url)
        prompt = (
            "请对该人脸进行情绪识别，并按以下格式输出：\n"
            "1) 先给出情绪类别（必须从以下集合中严格二选一）：\n"
            "【愤怒(Angry)、厌恶(Disgust)、恐惧(Fear)、高兴(Happy)、悲伤(Sad)、惊讶(Surprise)、中性(Neutral)】\n"
            "2) 然后给出简洁但更具体的中文分析（可包含面部特征线索，如眉眼、嘴角、皱纹、张口程度等）。\n"
            "请用中文回答，第一行仅输出类别（中文+括号内英文），从上述集合中选择其一；第二行开始给出简短分析。"
        )
        try:
            comp = client.chat.completions.create(
                model=model_name,
                messages=[{"role": "user", "content": [{"type": "image_url", "image_url": {"url": data_url}}, {"type": "text", "text": prompt}]}],
                timeout=60,
            )
        except Exception as exc:
            hint = ""
            try:
                # Provide actionable hints for common connection issues
                import ssl  # noqa: F401
            except Exception:
                pass
            proxy_hint = "已检测到网络连接异常。若处于代理/公司网络，请确保已设置 HTTP(S)_PROXY 环境变量，或临时关闭代理后再试。"
            hint = f"联网/代理问题：{proxy_hint}"
            raise RuntimeError(f"调用大模型失败：{exc}\n{hint}")
        # content can be a list of dicts or string depending on lib version
        content = comp.choices[0].message.content
        if isinstance(content, list):
            parts = []
            for p in content:
                if isinstance(p, dict) and p.get("type") == "text":
                    parts.append(p.get("text", ""))
            return "\n".join([t for t in parts if t])
        return str(content)

    def _analyze_worker(self) -> None:
        try:
            bgr = self._current_bgr.copy()
        except Exception:
            self._done_btn()
            return

        # Clear previous results table at start
        try:
            self.root.after(0, lambda: self.results_tv.delete(*self.results_tv.get_children()))
        except Exception:
            pass
        # Reset thumbnails store
        try:
            self._thumb_refs = []
        except Exception:
            pass

        t_task0 = time.perf_counter()
        dets = self._yolo_detect_faces(bgr)
        if not dets:
            # No faces, analyze the whole image
            dets = [{"box": [0, 0, bgr.shape[1]-1, bgr.shape[0]-1], "conf": 1.0}]

        # Prepare an annotated copy that shows only detection boxes (no emotion text)
        det_annotated = bgr.copy()
        model_name = self.model_name_var.get().strip() or "Qwen/Qwen2.5-VL-72B-Instruct"
        # reset export rows holder
        try:
            self._result_items = []
        except Exception:
            self._result_items = []
        # prepare thumbnail output directory for this analysis session
        try:
            ts_dir = datetime.now().strftime("%Y%m%d_%H%M%S")
            self._thumb_dir = (self.parent_app.output_dir / "emotion_thumbs" / ts_dir)
            self._thumb_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            self._thumb_dir = None

        # Determine a stable order: left-to-right, then top-to-bottom
        try:
            ordered = sorted(dets, key=lambda d: (int(d.get("box", [0, 0, 0, 0])[0]), int(d.get("box", [0, 0, 0, 0])[1])))
        except Exception:
            ordered = dets

        # Draw boxes with index labels first and update the main image immediately
        for seq, det in enumerate(ordered, start=1):
            x1, y1, x2, y2 = det["box"]
            x1 = max(0, x1); y1 = max(0, y1); x2 = min(bgr.shape[1]-1, x2); y2 = min(bgr.shape[0]-1, y2)
            cv2.rectangle(det_annotated, (x1, y1), (x2, y2), (0, 200, 255), 2)
            label_txt = str(seq)
            (tw, th), _ = cv2.getTextSize(label_txt, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)
            bx2 = x1 + tw + 8
            by2 = max(0, y1 - 6)
            by1 = max(0, by2 - th - 6)
            cv2.rectangle(det_annotated, (x1, by1), (bx2, by2), (0, 200, 255), -1)
            cv2.putText(det_annotated, label_txt, (x1 + 4, by2 - 4), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)

        # Show annotated image (with indices) before running LLM analysis
        self._last_annotated = det_annotated
        try:
            self.root.after(0, self._show_image, self._last_annotated)
        except Exception:
            pass

        # Now run analysis per face in the same order, inserting rows in ascending index
        for seq, det in enumerate(ordered, start=1):
            x1, y1, x2, y2 = det["box"]
            x1 = max(0, x1); y1 = max(0, y1); x2 = min(bgr.shape[1]-1, x2); y2 = min(bgr.shape[0]-1, y2)
            face = bgr[y1:y2, x1:x2]
            if face.size == 0:
                continue
            try:
                data_url = self._encode_bgr_to_data_url(face)
                content = self._call_qwen(model_name, data_url)
            except Exception as exc:
                content = f"分析失败: {exc}"

            # Build thumbnail for the crop only (no text overlay)
            try:
                face_vis = face.copy()
                rgb = cv2.cvtColor(face_vis, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(rgb)
                img.thumbnail((getattr(self, "thumb_size", 96), getattr(self, "thumb_size", 96)), Image.LANCZOS)
                imgtk = ImageTk.PhotoImage(image=img)
                self._thumb_refs.append(imgtk)
                # Save thumbnail to file for export
                thumb_path = None
                try:
                    if self._thumb_dir is not None:
                        thumb_path = self._thumb_dir / f"face_{seq}.png"
                        img.save(str(thumb_path), format="PNG")
                except Exception:
                    thumb_path = None
            except Exception:
                imgtk = None
                thumb_path = None

            def _insert_row(i=seq, bx=(x1, y1, x2, y2), txt=content, im=imgtk, tpath=thumb_path):
                try:
                    bbox_txt = f"[{bx[0]}, {bx[1]}, {bx[2]}, {bx[3]}]"
                    short_txt = (txt or "").strip()
                    try:
                        wrapped_lines = textwrap.wrap(short_txt, width=34)
                        short_txt = "\n".join(wrapped_lines[:15])
                    except Exception:
                        pass
                    unique_iid = f"face-{i}-{int(time.time()*1000)}"
                    self.results_tv.insert("", tk.END, iid=unique_iid, text="", image=im, values=(i, bbox_txt, short_txt))
                    self._result_items.append({"idx": i, "bbox": bbox_txt, "emotion": short_txt, "thumb_path": (str(tpath) if tpath else None)})
                except Exception:
                    pass
            try:
                self.root.after(0, _insert_row)
            except Exception:
                _insert_row()

        # Append history record with finished time (OS time)
        try:
            total_ms = (time.perf_counter() - t_task0) * 1000.0
            finished_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            img_path = (self.image_path_var.get() or "").strip()
            idx_val = self._hist_counter
            def _insert_hist():
                try:
                    self.history_tv2.insert("", tk.END, values=(idx_val, model_name, img_path, f"{total_ms:.1f}", finished_at))
                except Exception:
                    pass
            self.root.after(0, _insert_hist)
            self._hist_counter += 1
        finally:
            self._done_btn()

    def _done_btn(self) -> None:
        try:
            self.analyze_btn.config(state=tk.NORMAL)
        except Exception:
            pass

def main() -> None:
    root = tk.Tk()
    # Use system theme if available
    try:
        style = ttk.Style()
        if "vista" in style.theme_names():
            style.theme_use("vista")
        elif "clam" in style.theme_names():
            style.theme_use("clam")
    except Exception:
        pass

    app = FaceDetectionApp(root)
    root.protocol("WM_DELETE_WINDOW", lambda: (app.on_stop_camera(), root.destroy()))
    root.mainloop()


if __name__ == "__main__":
    main()


